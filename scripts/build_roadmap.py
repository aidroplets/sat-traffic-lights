"""
Builds ROADMAP.xlsx — the Study Signal feature roadmap.

Two sheets:
  • Roadmap — every feature, status, priority, difficulty, score
  • Legend — what each status / priority / difficulty means

Designed to be opened directly in Google Sheets (drag into Drive →
right-click → Open with Google Sheets, or File → Import).
"""
import datetime as dt
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle
)
from openpyxl.formatting.rule import (
    CellIsRule, FormulaRule, ColorScaleRule, DataBarRule
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.dimensions import ColumnDimension


# ============================================================
# Data
# ============================================================
# Each row: id, feature, description, category, status, priority,
# difficulty, effort_hrs, impact, owner, source, notes, date
# Score is a formula derived from impact + difficulty.

TODAY = dt.date(2026, 4, 29)

ROWS = [
    # ---- Shipped (already live) ------------------------------------
    ("F-001", "Score selection with clickable tiers",
     "Pick a target score (200–800); tier pills (≤500/501–600/601–700/701–800) are tappable shortcuts.",
     "Core", "Shipped", "P0", "S", 4, 4, "Joshua", "Spec PDF", "", TODAY),
    ("F-002", "Tiered question pools",
     "Test draws from difficulty buckets matching the target score.",
     "Core", "Shipped", "P0", "M", 4, 5, "Joshua", "Spec PDF", "", TODAY),
    ("F-003", "30-question test with G/Y/R signal",
     "Each question gets a green/yellow/red confidence signal — the core of the product.",
     "Core", "Shipped", "P0", "L", 8, 5, "Joshua", "Spec PDF", "", TODAY),
    ("F-004", "Live timer + total time",
     "Sticky-header timer during the test, total time on results.",
     "Core", "Shipped", "P1", "S", 3, 4, "Joshua", "Spec PDF", "", TODAY),
    ("F-005", "Inline error states (no alerts)",
     "Errors render in the UI rather than blocking modal dialogs.",
     "Core", "Shipped", "P1", "S", 3, 3, "Joshua", "Internal", "", TODAY),
    ("F-006", "Results bento (Mastery/Correct/Lucky/Review)",
     "Four-tile breakdown — true mastery vs. lucky vs. to-review.",
     "Core", "Shipped", "P0", "M", 4, 5, "Joshua", "Spec PDF", "", TODAY),
    ("F-007", "Walkthrough review with explanations",
     "Click-through review of every non-mastery question with the explanation.",
     "Core", "Shipped", "P0", "L", 8, 5, "Joshua", "Spec PDF", "", TODAY),
    ("F-008", "Share my score (image)",
     "Generates a 1080×1080 share-card PNG with score + breakdown.",
     "Growth", "Shipped", "P1", "M", 4, 4, "Joshua", "Spec PDF", "", TODAY),
    ("F-009", "Copy review link (URL)",
     "Encodes the attempt in a URL so you can send a review to a friend.",
     "Growth", "Shipped", "P1", "M", 4, 4, "Joshua", "Internal", "", TODAY),
    ("F-010", "Save trouble problems → PDF",
     "Compiles every non-mastery question with explanations into one printable PDF.",
     "Core", "Shipped", "P1", "M", 4, 4, "Joshua", "Spec PDF", "", TODAY),
    ("F-011", "History (past attempts)",
     "Persistent attempt history with open / re-share / delete / export.",
     "Core", "Shipped", "P1", "L", 8, 4, "Joshua", "Internal", "", TODAY),
    ("F-012", "Custom tailored test",
     "Re-test that drills topics from non-mastery items.",
     "Core", "Shipped", "P1", "M", 4, 4, "Joshua", "Spec PDF", "", TODAY),
    ("F-013", "Admin dashboard with filters + actions",
     "Wide table view of every question — filter by state/source/topic, Publish/Unpublish/Edit per row.",
     "Admin", "Shipped", "P0", "L", 10, 5, "Joshua", "Internal", "", TODAY),
    ("F-014", "Question schema + provenance",
     "state (live/unpublished/needs-review), source (human-curated/ai-generated), uploader email.",
     "Content", "Shipped", "P0", "M", 4, 5, "Joshua", "Internal", "", TODAY),
    ("F-015", "100 AI-generated questions (needs-review)",
     "Pilot batch in needs-review state, awaiting admin audit before going live.",
     "Content", "Shipped", "P1", "L", 6, 4, "Joshua", "Internal", "", TODAY),
    ("F-016", "Magic-link login (email)",
     "Email-only auth via Resend; HttpOnly signed JWT cookies; 30-day sessions.",
     "Auth", "Shipped", "P0", "L", 8, 5, "Joshua", "Internal", "", TODAY),
    ("F-017", "Roles + admin gating",
     "user vs. admin roles. Admin allowlist for joshua@sortino.co + mweber0204@gmail.com.",
     "Auth", "Shipped", "P0", "M", 4, 5, "Joshua", "Internal", "", TODAY),
    ("F-018", "Resend integration + verified domain",
     "noreply@studysignal.ai sends magic-link emails; rate-limited per email.",
     "Infra", "Shipped", "P0", "M", 3, 4, "Joshua", "Internal", "", TODAY),
    ("F-019", "Mobile-first responsive design",
     "Safe-area insets, touch baseline, hover-on-touch override, breakpoint ladder.",
     "Core", "Shipped", "P0", "L", 8, 5, "Joshua", "Internal", "", TODAY),
    ("F-020", "Auth UI in distinct aesthetic",
     "Compact centered card with halo glow — visually distinct from test cards.",
     "Auth", "Shipped", "P1", "M", 3, 3, "Joshua", "Internal", "", TODAY),
    ("F-021", "Ship to studysignal.ai (custom domain)",
     "Production deploy, DNS at Hover, SSL via Let's Encrypt.",
     "Infra", "Shipped", "P0", "M", 3, 4, "Joshua", "Internal", "", TODAY),

    # ---- Next up (committed, near-term) ----------------------------
    ("F-022", "Audit the 100 AI questions",
     "Manual review pass: for each AI-generated question, verify answer key, promote to live.",
     "Content", "Next up", "P0", "L", 10, 5, "Joshua", "Internal",
     "Filter admin → needs-review + ai-generated, work through them.", TODAY),
    ("F-023", "Supabase database setup",
     "Replace in-memory dev store with Postgres. Run db/001_initial_schema.sql.",
     "Infra", "Next up", "P0", "M", 3, 5, "Joshua", "Internal",
     "Steps documented in BACKEND-SETUP.md.", TODAY),
    ("F-024", "Cross-device attempt sync",
     "Save attempts to Supabase so history follows the user across browsers/devices.",
     "Infra", "Next up", "P1", "L", 6, 4, "", "Internal",
     "Depends on Supabase being live (F-023).", TODAY),
    ("F-025", "Admin allowlist via DB (not code)",
     "Move ADMIN_EMAILS Set out of source. Admin grants happen via SQL update.",
     "Admin", "Next up", "P2", "S", 2, 3, "", "Internal",
     "Depends on F-023.", TODAY),
    ("F-026", "Sharper 'Similar problems' tailored test",
     "Broaden pool to all non-(green-and-correct) items; pick non-identical questions on same topic+difficulty.",
     "Core", "Next up", "P1", "M", 4, 4, "", "Spec PDF", "", TODAY),
    ("F-027", "Challenge a friend (send a test)",
     "Encode a fixed test setup in a URL; friend takes the same questions; results comparison.",
     "Growth", "Next up", "P1", "M", 5, 4, "", "Spec PDF", "", TODAY),
    ("F-028", "Public signup flow polish",
     "Already works; needs onboarding-friendly copy + first-test handholding for non-admin users.",
     "Auth", "Next up", "P1", "S", 2, 4, "", "Internal", "", TODAY),

    # ---- Backlog (known-want, scheduled later) ---------------------
    ("F-029", "Generate 1000 more AI questions",
     "Continue batches of 100 with the same self-validation methodology. Full bank target ~1100.",
     "Content", "Backlog", "P1", "XL", 30, 5, "", "User feedback",
     "Wait for F-022 to validate quality bar first.", TODAY),
    ("F-030", "Reading + writing sections",
     "Currently math-only. Add the two verbal sections of the SAT.",
     "Content", "Backlog", "P2", "XL", 40, 5, "", "Internal", "", TODAY),
    ("F-031", "ACT mode (separate test format)",
     "ACT scoring + question pool + section structure.",
     "Core", "Backlog", "P2", "L", 16, 4, "", "Internal", "", TODAY),
    ("F-032", "Practice mode (no timer, no scoring)",
     "Drill mode for low-pressure repetition.",
     "Core", "Backlog", "P2", "S", 3, 3, "", "Internal", "", TODAY),
    ("F-033", "Adaptive difficulty",
     "Adjust the pool mid-test based on early-question signals.",
     "Core", "Backlog", "P2", "XL", 30, 4, "", "Internal", "", TODAY),
    ("F-034", "Daily streak / habit tracking",
     "Streak counter, daily reminder, calendar heatmap.",
     "Growth", "Backlog", "P2", "M", 5, 3, "", "Internal", "", TODAY),
    ("F-035", "Weekly email digest",
     "Personalized rundown of weak topics + recommended drills.",
     "Growth", "Backlog", "P2", "M", 6, 3, "", "Internal", "", TODAY),
    ("F-036", "Progress visualization over time",
     "Score trajectory, mastery-by-topic heatmap, time-spent.",
     "Core", "Backlog", "P2", "L", 10, 4, "", "Internal", "", TODAY),
    ("F-037", "Question favoriting + notes",
     "Star a question; attach a personal note; review favorites later.",
     "Core", "Backlog", "P3", "S", 3, 2, "", "Internal", "", TODAY),
    ("F-038", "Spaced repetition for missed",
     "Re-surface missed questions on Leitner-box schedule.",
     "Core", "Backlog", "P2", "L", 12, 4, "", "Internal", "", TODAY),
    ("F-039", "PWA + offline support",
     "Service worker + manifest. Already responsive — needs caching strategy.",
     "Infra", "Backlog", "P2", "L", 10, 3, "", "Internal", "", TODAY),
    ("F-040", "Onboarding tutorial",
     "First-run walkthrough explaining the green/yellow/red mechanic.",
     "Growth", "Backlog", "P2", "M", 5, 3, "", "Internal", "", TODAY),
    ("F-041", "Marketing site",
     "Public-facing landing page with positioning, screenshots, pricing.",
     "Growth", "Backlog", "P3", "L", 16, 3, "", "Internal", "", TODAY),
    ("F-042", "Stripe integration (paid tier)",
     "Free tier vs. paid (e.g., unlimited tests, advanced stats).",
     "Infra", "Backlog", "P3", "L", 14, 4, "", "Internal", "", TODAY),
    ("F-043", "Question stats dashboard",
     "Admin: cohort-level stats per question — % correct, signal distribution, weak topics.",
     "Admin", "Backlog", "P2", "M", 5, 3, "", "Internal", "", TODAY),
    ("F-044", "Bulk import questions",
     "Admin: paste/upload CSV or JSON to add questions in batch.",
     "Admin", "Backlog", "P2", "M", 4, 3, "", "Internal", "", TODAY),
    ("F-045", "Subject-specific scoring breakdown",
     "Show per-section scoring estimate (math vs. reading vs. writing) on results.",
     "Core", "Backlog", "P2", "M", 4, 3, "", "Internal",
     "Depends on F-030.", TODAY),
    ("F-046", "A/B testing framework",
     "Feature-flag system + cohort assignment for product experiments.",
     "Infra", "Backlog", "P3", "L", 14, 3, "", "Internal", "", TODAY),

    # ---- Ideas (speculative, longer-term) --------------------------
    ("F-047", "Comparison to peers",
     "Cohort percentile + benchmark scores from anonymized aggregate.",
     "Future", "Idea", "P3", "L", 14, 3, "", "Internal", "", TODAY),
    ("F-048", "Tutor mode (AI help)",
     "Embedded AI tutor — explain stuck problems, generate hints.",
     "Future", "Idea", "P3", "XL", 40, 4, "", "Internal", "", TODAY),
    ("F-049", "Class / teacher dashboard",
     "Teachers create cohorts, assign tests, see aggregate progress.",
     "Future", "Idea", "P3", "XL", 50, 4, "", "Internal", "", TODAY),
    ("F-050", "Native mobile app",
     "iOS + Android wrappers — push notifications, deep links, App Store presence.",
     "Future", "Idea", "P3", "XL", 80, 3, "", "Internal",
     "PWA (F-039) likely covers 80% of value at 10% of cost.", TODAY),
]

HEADERS = [
    "ID", "Feature", "Description", "Category", "Status", "Priority",
    "Difficulty", "Effort (hrs)", "Impact", "Score", "Owner",
    "Source", "Notes", "Date added"
]


# ============================================================
# Color palette — matches the Study Signal app aesthetic
# ============================================================
NAVY      = "0d0e18"     # dark canvas, header
LAVENDER  = "8b86ff"     # accent
LAV_SOFT  = "1f1d3a"     # soft accent
HAIRLINE  = "3a3d4a"
TEXT      = "f0f2f7"
MUTED     = "a8b0c4"

# status fills
S_SHIPPED  = "1a4d2e"    # dark green
S_NEXT     = "1a3a5c"    # dark blue
S_BACKLOG  = "3a3d4a"    # neutral gray
S_IDEA     = "2a2030"    # subtle purple
S_BLOCKED  = "5c1a1a"    # dark red

# priority fills (subtler)
P0 = "5c1a1a"
P1 = "5c3e1a"
P2 = "3a3d4a"
P3 = "2a2a30"

# zebra striping
BAND_A = "12131c"
BAND_B = "0d0e18"


# ============================================================
# Build workbook
# ============================================================

wb = Workbook()
ws = wb.active
ws.title = "Roadmap"

# ---- Header row ----
header_font = Font(name="Arial", bold=True, size=11, color=TEXT)
header_fill = PatternFill("solid", start_color=LAVENDER, end_color=LAVENDER)
header_align = Alignment(horizontal="left", vertical="center")

ws.append(HEADERS)
for cell in ws[1]:
    cell.font = Font(name="Arial", bold=True, size=11, color="0d0e18")
    cell.fill = PatternFill("solid", start_color=LAVENDER)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = Border(bottom=Side(style="thin", color=NAVY))
ws.row_dimensions[1].height = 32

# ---- Data rows ----
data_font = Font(name="Arial", size=10, color=TEXT)
band_a_fill = PatternFill("solid", start_color=BAND_A)
band_b_fill = PatternFill("solid", start_color=BAND_B)

for row_idx, row in enumerate(ROWS, start=2):
    (rid, feat, desc, cat, status, prio, diff, hrs, impact, owner,
     src, notes, date) = row
    excel_row = [
        rid, feat, desc, cat, status, prio, diff, hrs, impact,
        # Score formula: Impact × difficulty-bonus, where XS=6, S=5, M=4, L=3, XL=2.
        # Higher impact + lower effort = higher score.
        f'=IFERROR(I{row_idx}*CHOOSE(MATCH(G{row_idx},{{"XS","S","M","L","XL"}},0),6,5,4,3,2),"")',
        owner, src, notes, date,
    ]
    ws.append(excel_row)

    fill = band_a_fill if row_idx % 2 == 0 else band_b_fill
    for cell in ws[row_idx]:
        cell.font = data_font
        cell.fill = fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.row_dimensions[row_idx].height = 38

# ---- Column widths ----
widths = {
    "A": 8,    # ID
    "B": 38,   # Feature
    "C": 56,   # Description
    "D": 12,   # Category
    "E": 12,   # Status
    "F": 9,    # Priority
    "G": 10,   # Difficulty
    "H": 11,   # Effort
    "I": 9,    # Impact
    "J": 9,    # Score
    "K": 12,   # Owner
    "L": 14,   # Source
    "M": 36,   # Notes
    "N": 13,   # Date added
}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# ---- Number/date format ----
for r in range(2, len(ROWS) + 2):
    ws.cell(row=r, column=10).number_format = "0.0"           # Score
    ws.cell(row=r, column=14).number_format = "yyyy-mm-dd"    # Date

# ---- Freeze + auto-filter ----
ws.freeze_panes = "C2"
ws.auto_filter.ref = ws.dimensions

# ---- Conditional formatting: status column (E) ----
last_row = len(ROWS) + 1

def fill_rule(text, fg, bg):
    return CellIsRule(
        operator="equal",
        formula=[f'"{text}"'],
        stopIfTrue=False,
        fill=PatternFill("solid", start_color=bg),
        font=Font(name="Arial", size=10, color=fg, bold=True),
    )

ws.conditional_formatting.add(f"E2:E{last_row}", fill_rule("Shipped",  TEXT, S_SHIPPED))
ws.conditional_formatting.add(f"E2:E{last_row}", fill_rule("Next up",  TEXT, S_NEXT))
ws.conditional_formatting.add(f"E2:E{last_row}", fill_rule("Backlog",  MUTED, S_BACKLOG))
ws.conditional_formatting.add(f"E2:E{last_row}", fill_rule("Idea",     MUTED, S_IDEA))
ws.conditional_formatting.add(f"E2:E{last_row}", fill_rule("Blocked",  TEXT, S_BLOCKED))

# ---- Conditional formatting: priority column (F) ----
ws.conditional_formatting.add(f"F2:F{last_row}", fill_rule("P0", TEXT,  P0))
ws.conditional_formatting.add(f"F2:F{last_row}", fill_rule("P1", TEXT,  P1))
ws.conditional_formatting.add(f"F2:F{last_row}", fill_rule("P2", MUTED, P2))
ws.conditional_formatting.add(f"F2:F{last_row}", fill_rule("P3", MUTED, P3))

# ---- Data bar on Score (J) — visual ranking ----
ws.conditional_formatting.add(
    f"J2:J{last_row}",
    DataBarRule(
        start_type="min", end_type="max",
        color=LAVENDER, showValue=True,
    )
)

# ============================================================
# Sheet 2 — Legend
# ============================================================
legend = wb.create_sheet("Legend")

LEGEND_BLOCKS = [
    ("Status", [
        ("Shipped",  "Built and live on studysignal.ai."),
        ("Next up",  "Committed for the next batch of work."),
        ("Backlog",  "Known-want, no immediate timeline."),
        ("Idea",     "Speculative — not committed."),
        ("Blocked",  "Needs a decision or upstream work."),
    ]),
    ("Priority", [
        ("P0", "Critical — ship immediately."),
        ("P1", "High — within next batch."),
        ("P2", "Medium — known-want, no urgency."),
        ("P3", "Low — nice to have."),
    ]),
    ("Difficulty", [
        ("XS", "Less than 1 hour."),
        ("S",  "Less than 4 hours."),
        ("M",  "Less than 1 day."),
        ("L",  "Multi-day, single chunk of work."),
        ("XL", "A week or more, often multiple parts."),
    ]),
    ("Category", [
        ("Core",    "The actual study experience."),
        ("Content", "Question pipeline + bank."),
        ("Admin",   "Internal admin tooling."),
        ("Auth",    "Sign-in, accounts, roles."),
        ("Infra",   "Backend, deployment, ops."),
        ("Growth",  "Marketing, sharing, retention."),
        ("Future",  "Speculative, longer-term."),
    ]),
    ("Score", [
        ("Formula", "Impact × difficulty-bonus (XS=6 → XL=2). Higher = ship-sooner candidate."),
        ("Range",   "1–30. Use as a sort key, not a hard rule."),
    ]),
]

# Title row
legend["A1"] = "Roadmap legend"
legend["A1"].font = Font(name="Arial", bold=True, size=14, color=TEXT)
legend["A1"].fill = PatternFill("solid", start_color=LAVENDER)
legend["A1"].alignment = Alignment(vertical="center")
legend.merge_cells("A1:B1")
legend.row_dimensions[1].height = 30

row = 3
for block_title, items in LEGEND_BLOCKS:
    legend.cell(row=row, column=1, value=block_title).font = Font(name="Arial", bold=True, size=11, color=LAVENDER)
    legend.cell(row=row, column=1).fill = PatternFill("solid", start_color=LAV_SOFT)
    legend.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    legend.row_dimensions[row].height = 22
    row += 1
    for k, v in items:
        c1 = legend.cell(row=row, column=1, value=k)
        c2 = legend.cell(row=row, column=2, value=v)
        c1.font = Font(name="Arial", bold=True, size=10, color=TEXT)
        c2.font = Font(name="Arial", size=10, color=MUTED)
        c1.alignment = Alignment(vertical="top")
        c2.alignment = Alignment(vertical="top", wrap_text=True)
        c1.fill = PatternFill("solid", start_color=BAND_A)
        c2.fill = PatternFill("solid", start_color=BAND_A)
        legend.row_dimensions[row].height = 22
        row += 1
    row += 1   # blank row between blocks

legend.column_dimensions["A"].width = 16
legend.column_dimensions["B"].width = 80

# ============================================================
# Save
# ============================================================
out = "/sessions/wizardly-epic-pascal/mnt/aidroplets.com/droplets/sat-traffic-lights/ROADMAP.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Rows: {len(ROWS)}")
